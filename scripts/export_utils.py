#!/usr/bin/env python3
"""
Export Utils - Export výsledkov do CSV a Excel s grafmi

Funkcie:
1. Export do CSV - jednoduchý formát
2. Export do Excel s formátovaním a grafmi
"""
import csv
import os
from datetime import datetime
from typing import Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, Reference, BarChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False


class ExportUtils:
    """
    Utility trieda pre export dát do rôznych formátov
    """
    
    # Farby pre zóny P/L
    COLORS = {
        'profit': '90EE90',      # Light green
        'loss': 'FFB6C1',        # Light red
        'neutral': 'FFFACD',     # Light yellow
        'header': '4472C4',      # Blue
        'header_font': 'FFFFFF', # White
    }
    
    @staticmethod
    def generate_filename(symbol: str, prefix: str = 'hedge_analysis') -> str:
        """
        Generuje unikátny názov súboru
        
        Args:
            symbol: Symbol (napr. SPY)
            prefix: Prefix názvu súboru
            
        Returns:
            Názov súboru s dátumom
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{prefix}_{symbol}_{timestamp}"
    
    @staticmethod
    def export_to_csv(data: Dict, filepath: str) -> str:
        """
        Export stratégie a scenárov do CSV
        
        Args:
            data: Dict s dátami na export
            filepath: Cesta k výstupnému súboru (bez prípony)
            
        Returns:
            Cesta k vytvorenému súboru
        """
        csv_path = f"{filepath}.csv"
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # === SEKCIA 1: Prehľad stratégie ===
            writer.writerow(['=== HEDGE STRATEGY OVERVIEW ==='])
            writer.writerow([])
            
            if 'symbol' in data:
                writer.writerow(['Symbol', data.get('symbol', '')])
                writer.writerow(['Current Price', data.get('currentPrice', '')])
                writer.writerow(['Option Type', data.get('optionType', '')])
            
            # Short leg
            if 'shortLeg' in data:
                writer.writerow([])
                writer.writerow(['=== SHORT LEG ==='])
                short = data['shortLeg']
                writer.writerow(['Strike', short.get('strike', '')])
                writer.writerow(['Expiry', short.get('expiry', '')])
                writer.writerow(['Premium', short.get('premium', '')])
                writer.writerow(['Delta', short.get('delta', '')])
                writer.writerow(['Theta', short.get('theta', '')])
                writer.writerow(['IV', short.get('iv', '')])
            
            # Long leg
            if 'longLeg' in data:
                writer.writerow([])
                writer.writerow(['=== LONG LEG ==='])
                long = data['longLeg']
                writer.writerow(['Strike', long.get('strike', '')])
                writer.writerow(['Expiry', long.get('expiry', '')])
                writer.writerow(['Premium', long.get('premium', '')])
                writer.writerow(['Delta', long.get('delta', '')])
                writer.writerow(['Theta', long.get('theta', '')])
                writer.writerow(['IV', long.get('iv', '')])
            
            # Strategy summary
            if 'strategy' in data:
                writer.writerow([])
                writer.writerow(['=== STRATEGY SUMMARY ==='])
                strat = data['strategy']
                writer.writerow(['Net Credit', strat.get('netCredit', '')])
                writer.writerow(['Max Profit', strat.get('maxProfit', '')])
                writer.writerow(['Max Loss', strat.get('maxLoss', '')])
                writer.writerow(['Breakeven', strat.get('breakeven', '')])
                writer.writerow(['Spread Width', strat.get('spreadWidth', '')])
                writer.writerow(['Margin Required', strat.get('marginRequired', '')])
            
            # Margin info
            if 'marginInfo' in data:
                writer.writerow([])
                writer.writerow(['=== MARGIN INFO ==='])
                margin = data['marginInfo']
                writer.writerow(['Broker', margin.get('broker', '')])
                writer.writerow(['Spread Type', margin.get('spreadType', '')])
                writer.writerow(['Margin', margin.get('margin', '')])
                writer.writerow(['ROI on Margin', f"{margin.get('roiOnMargin', '')}%"])
                writer.writerow(['Weekly ROI', f"{margin.get('weeklyROI', '')}%"])
            
            # Alternatives
            if 'alternatives' in data:
                writer.writerow([])
                writer.writerow(['=== ALTERNATIVES (by Weekly ROI) ==='])
                writer.writerow(['DTE Offset', 'Long Strike', 'Margin', 'Net Credit', 
                               'Weekly ROI %', 'Theta Adjusted ROI %'])
                for alt in data['alternatives']:
                    writer.writerow([
                        alt.get('dteOffset', ''),
                        alt.get('longStrike', ''),
                        alt.get('margin', ''),
                        alt.get('netCredit', ''),
                        alt.get('weeklyROI', ''),
                        alt.get('thetaAdjustedWeeklyROI', ''),
                    ])
            
            # Scenarios - Price moves
            if 'priceScenarios' in data:
                writer.writerow([])
                writer.writerow(['=== PRICE MOVE SCENARIOS ==='])
                writer.writerow(['Price Change %', 'New Price', 'P/L USD', 'Net Delta'])
                for s in data['priceScenarios']:
                    writer.writerow([
                        s.get('priceChange', ''),
                        s.get('newPrice', ''),
                        s.get('pnl', ''),
                        s.get('netDelta', ''),
                    ])
            
            # Scenarios - Time decay
            if 'timeScenarios' in data:
                writer.writerow([])
                writer.writerow(['=== TIME DECAY SCENARIOS ==='])
                writer.writerow(['Days Forward', 'Short DTE', 'Long DTE', 'P/L USD'])
                for s in data['timeScenarios']:
                    writer.writerow([
                        s.get('daysForward', ''),
                        s.get('shortDTE', ''),
                        s.get('longDTE', ''),
                        s.get('pnl', ''),
                    ])
            
            # Combined P/L matrix
            if 'combinedMatrix' in data:
                writer.writerow([])
                writer.writerow(['=== P/L MATRIX (Price x Time) ==='])
                
                matrix = data['combinedMatrix']
                price_changes = matrix.get('priceChanges', [])
                
                # Header row
                header = ['DTE'] + [f"{p:+.0f}%" for p in price_changes]
                writer.writerow(header)
                
                # Data rows
                for row in matrix.get('matrix', []):
                    csv_row = [row.get('shortDTE', '')]
                    for scenario in row.get('scenarios', []):
                        csv_row.append(scenario.get('pnl', ''))
                    writer.writerow(csv_row)
        
        return csv_path
    
    @staticmethod
    def export_to_excel(data: Dict, filepath: str) -> str:
        """
        Export stratégie a scenárov do Excel s formátovaním a grafmi
        
        Args:
            data: Dict s dátami na export
            filepath: Cesta k výstupnému súboru (bez prípony)
            
        Returns:
            Cesta k vytvorenému súboru
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl nie je nainštalovaný. Použite: pip install openpyxl")
        
        xlsx_path = f"{filepath}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # === LIST 1: Prehľad stratégie ===
        ws1 = wb.active
        ws1.title = "Prehľad"
        ExportUtils._write_overview_sheet(ws1, data)
        
        # === LIST 2: Scenárová analýza ===
        ws2 = wb.create_sheet("Scenáre")
        ExportUtils._write_scenarios_sheet(ws2, data)
        
        # === LIST 3: Alternatívy ===
        if 'alternatives' in data:
            ws3 = wb.create_sheet("Alternatívy")
            ExportUtils._write_alternatives_sheet(ws3, data)
        
        # === LIST 4: P/L Matrix ===
        if 'combinedMatrix' in data:
            ws4 = wb.create_sheet("P-L Matrix")
            ExportUtils._write_matrix_sheet(ws4, data)
        
        wb.save(xlsx_path)
        return xlsx_path
    
    @staticmethod
    def _write_overview_sheet(ws, data: Dict):
        """Zapíše prehľadový list"""
        # Štýly
        header_font = Font(bold=True, color=ExportUtils.COLORS['header_font'])
        header_fill = PatternFill(start_color=ExportUtils.COLORS['header'], 
                                  end_color=ExportUtils.COLORS['header'], 
                                  fill_type='solid')
        bold_font = Font(bold=True)
        
        row = 1
        
        # Hlavička
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"HEDGE ANALYSIS - {data.get('symbol', 'N/A')}"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Info o symbole
        ws[f'A{row}'] = "Symbol:"
        ws[f'B{row}'] = data.get('symbol', '')
        ws[f'A{row}'].font = bold_font
        row += 1
        
        ws[f'A{row}'] = "Aktuálna cena:"
        ws[f'B{row}'] = f"${data.get('currentPrice', 0):.2f}"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        ws[f'A{row}'] = "Typ opcie:"
        ws[f'B{row}'] = data.get('optionType', 'PUT')
        ws[f'A{row}'].font = bold_font
        row += 2
        
        # Short leg
        ws[f'A{row}'] = "SHORT LEG"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        short = data.get('shortLeg', {})
        for key, val in [('Strike', short.get('strike')), 
                         ('Expirácia', short.get('expiry')),
                         ('Premium', f"${short.get('premium', 0):.2f}" if short.get('premium') else ''),
                         ('Delta', f"{short.get('delta', 0):.4f}" if short.get('delta') else ''),
                         ('Theta', f"{short.get('theta', 0):.4f}" if short.get('theta') else ''),
                         ('IV', f"{short.get('iv', 0)*100:.1f}%" if short.get('iv') else '')]:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = val
            row += 1
        
        row += 1
        
        # Long leg
        ws[f'A{row}'] = "LONG LEG"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        long = data.get('longLeg', {})
        for key, val in [('Strike', long.get('strike')), 
                         ('Expirácia', long.get('expiry')),
                         ('Premium', f"${long.get('premium', 0):.2f}" if long.get('premium') else ''),
                         ('Delta', f"{long.get('delta', 0):.4f}" if long.get('delta') else ''),
                         ('Theta', f"{long.get('theta', 0):.4f}" if long.get('theta') else '')]:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = val
            row += 1
        
        row += 1
        
        # Strategy summary
        ws[f'A{row}'] = "STRATÉGIA"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        strat = data.get('strategy', {})
        for key, val in [('Net Credit', f"${strat.get('netCredit', 0):.2f}"),
                         ('Max Profit', f"${strat.get('maxProfit', 0):.2f}"),
                         ('Max Loss', f"${strat.get('maxLoss', 0):.2f}"),
                         ('Breakeven', f"${strat.get('breakeven', 0):.2f}"),
                         ('Spread Width', strat.get('spreadWidth', '')),
                         ('Margin', f"${strat.get('marginRequired', 0):.2f}")]:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = val
            row += 1
        
        # Margin info
        if 'marginInfo' in data:
            row += 1
            ws[f'A{row}'] = "MARGIN INFO"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            margin = data['marginInfo']
            for key, val in [('Broker', margin.get('brokerName', margin.get('broker', ''))),
                             ('Typ spreadu', margin.get('spreadType', '')),
                             ('Margin', f"${margin.get('margin', 0):.2f}"),
                             ('ROI on Margin', f"{margin.get('roiOnMargin', 0):.2f}%"),
                             ('Weekly ROI', f"{margin.get('weeklyROI', 0):.2f}%"),
                             ('Theta Adj. ROI', f"{margin.get('thetaAdjustedWeeklyROI', 0):.2f}%")]:
                ws[f'A{row}'] = key
                ws[f'B{row}'] = val
                row += 1
        
        # Šírka stĺpcov
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
    
    @staticmethod
    def _write_scenarios_sheet(ws, data: Dict):
        """Zapíše scenárovú analýzu s grafom"""
        header_font = Font(bold=True, color=ExportUtils.COLORS['header_font'])
        header_fill = PatternFill(start_color=ExportUtils.COLORS['header'], 
                                  end_color=ExportUtils.COLORS['header'], 
                                  fill_type='solid')
        
        row = 1
        
        # === Price scenarios ===
        if 'priceScenarios' in data:
            ws[f'A{row}'] = "SCENÁRE - POHYB CENY"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            headers = ['Zmena ceny %', 'Nová cena', 'P/L USD', 'Net Delta']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            price_data_start = row
            for s in data['priceScenarios']:
                ws.cell(row=row, column=1, value=s.get('priceChange', 0))
                ws.cell(row=row, column=2, value=s.get('newPrice', 0))
                ws.cell(row=row, column=3, value=s.get('pnl', 0))
                ws.cell(row=row, column=4, value=s.get('netDelta', 0))
                
                # Farebné zvýraznenie P/L
                pnl = s.get('pnl', 0)
                if pnl > 0:
                    ws.cell(row=row, column=3).fill = PatternFill(
                        start_color=ExportUtils.COLORS['profit'],
                        end_color=ExportUtils.COLORS['profit'],
                        fill_type='solid'
                    )
                elif pnl < 0:
                    ws.cell(row=row, column=3).fill = PatternFill(
                        start_color=ExportUtils.COLORS['loss'],
                        end_color=ExportUtils.COLORS['loss'],
                        fill_type='solid'
                    )
                row += 1
            price_data_end = row - 1
            
            # Graf pre price scenarios
            if price_data_end > price_data_start:
                chart = LineChart()
                chart.title = "P/L vs Price Change"
                chart.y_axis.title = "P/L (USD)"
                chart.x_axis.title = "Price Change %"
                
                data_ref = Reference(ws, min_col=3, min_row=price_data_start-1, 
                                    max_row=price_data_end, max_col=3)
                cats = Reference(ws, min_col=1, min_row=price_data_start, 
                               max_row=price_data_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.width = 12
                chart.height = 8
                
                ws.add_chart(chart, "F2")
            
            row += 2
        
        # === Time decay scenarios ===
        if 'timeScenarios' in data:
            ws[f'A{row}'] = "SCENÁRE - ČASOVÝ ROZPAD"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            headers = ['Dni', 'Short DTE', 'Long DTE', 'P/L USD']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            time_data_start = row
            for s in data['timeScenarios']:
                ws.cell(row=row, column=1, value=s.get('daysForward', 0))
                ws.cell(row=row, column=2, value=s.get('shortDTE', 0))
                ws.cell(row=row, column=3, value=s.get('longDTE', 0))
                ws.cell(row=row, column=4, value=s.get('pnl', 0))
                
                pnl = s.get('pnl', 0)
                if pnl > 0:
                    ws.cell(row=row, column=4).fill = PatternFill(
                        start_color=ExportUtils.COLORS['profit'],
                        end_color=ExportUtils.COLORS['profit'],
                        fill_type='solid'
                    )
                row += 1
            time_data_end = row - 1
            
            # Graf pre time decay
            if time_data_end > time_data_start:
                chart2 = LineChart()
                chart2.title = "P/L vs Time (Theta Decay)"
                chart2.y_axis.title = "P/L (USD)"
                chart2.x_axis.title = "Days Forward"
                
                data_ref = Reference(ws, min_col=4, min_row=time_data_start-1,
                                    max_row=time_data_end, max_col=4)
                cats = Reference(ws, min_col=1, min_row=time_data_start,
                               max_row=time_data_end)
                chart2.add_data(data_ref, titles_from_data=True)
                chart2.set_categories(cats)
                chart2.width = 12
                chart2.height = 8
                
                ws.add_chart(chart2, "F15")
        
        # Šírka stĺpcov
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 14
    
    @staticmethod
    def _write_alternatives_sheet(ws, data: Dict):
        """Zapíše alternatívy s rôznymi DTE offsetmi"""
        header_font = Font(bold=True, color=ExportUtils.COLORS['header_font'])
        header_fill = PatternFill(start_color=ExportUtils.COLORS['header'], 
                                  end_color=ExportUtils.COLORS['header'], 
                                  fill_type='solid')
        
        ws['A1'] = "ALTERNATÍVY PODĽA DTE OFFSET"
        ws['A1'].font = Font(bold=True, size=12)
        
        headers = ['DTE Offset', 'Long Strike', 'Margin ($)', 'Net Credit ($)', 
                   'Weekly ROI %', 'Theta Adj. ROI %', 'Theta Diff']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 3
        for alt in data.get('alternatives', []):
            ws.cell(row=row, column=1, value=alt.get('dteOffset', 0))
            ws.cell(row=row, column=2, value=alt.get('longStrike', 0))
            ws.cell(row=row, column=3, value=alt.get('margin', 0))
            ws.cell(row=row, column=4, value=alt.get('netCredit', 0))
            ws.cell(row=row, column=5, value=alt.get('weeklyROI', 0))
            ws.cell(row=row, column=6, value=alt.get('thetaAdjustedWeeklyROI', 0))
            ws.cell(row=row, column=7, value=alt.get('thetaDifferential', 0))
            row += 1
        
        # Graf ROI vs DTE Offset
        if row > 3:
            chart = BarChart()
            chart.title = "Weekly ROI vs DTE Offset"
            chart.y_axis.title = "Weekly ROI %"
            chart.x_axis.title = "DTE Offset"
            
            data_ref = Reference(ws, min_col=5, min_row=2, max_row=row-1, max_col=6)
            cats = Reference(ws, min_col=1, min_row=3, max_row=row-1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 14
            chart.height = 8
            
            ws.add_chart(chart, "I2")
        
        # Šírka stĺpcov
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 14
    
    @staticmethod
    def _write_matrix_sheet(ws, data: Dict):
        """Zapíše P/L maticu s farebnými zónami"""
        header_font = Font(bold=True, color=ExportUtils.COLORS['header_font'])
        header_fill = PatternFill(start_color=ExportUtils.COLORS['header'],
                                  end_color=ExportUtils.COLORS['header'],
                                  fill_type='solid')
        
        matrix = data.get('combinedMatrix', {})
        price_changes = matrix.get('priceChanges', [])
        matrix_data = matrix.get('matrix', [])
        
        ws['A1'] = "P/L MATICA (Cena × Čas)"
        ws['A1'].font = Font(bold=True, size=12)
        
        ws['A2'] = f"Aktuálna cena: ${matrix.get('currentPrice', 0):.2f}"
        ws['A3'] = f"Net Credit: ${matrix.get('originalNetCredit', 0):.2f}"
        
        # Header row
        row = 5
        ws.cell(row=row, column=1, value="DTE").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        for col, pct in enumerate(price_changes, 2):
            cell = ws.cell(row=row, column=col, value=f"{pct:+.0f}%")
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row = 6
        for matrix_row in matrix_data:
            ws.cell(row=row, column=1, value=matrix_row.get('shortDTE', 0))
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for col, scenario in enumerate(matrix_row.get('scenarios', []), 2):
                cell = ws.cell(row=row, column=col, value=scenario.get('pnl', 0))
                
                # Farebné zvýraznenie podľa zóny
                zone = scenario.get('zone', 'neutral')
                if zone == 'profit':
                    cell.fill = PatternFill(start_color=ExportUtils.COLORS['profit'],
                                           end_color=ExportUtils.COLORS['profit'],
                                           fill_type='solid')
                elif zone == 'loss':
                    cell.fill = PatternFill(start_color=ExportUtils.COLORS['loss'],
                                           end_color=ExportUtils.COLORS['loss'],
                                           fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=ExportUtils.COLORS['neutral'],
                                           end_color=ExportUtils.COLORS['neutral'],
                                           fill_type='solid')
            row += 1
        
        # Legenda
        row += 2
        ws.cell(row=row, column=1, value="Legenda:").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Zelená = Profit zóna (≥50% max profit)")
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color=ExportUtils.COLORS['profit'],
            end_color=ExportUtils.COLORS['profit'],
            fill_type='solid'
        )
        row += 1
        ws.cell(row=row, column=1, value="Červená = Loss zóna (≥50% max loss)")
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color=ExportUtils.COLORS['loss'],
            end_color=ExportUtils.COLORS['loss'],
            fill_type='solid'
        )
        row += 1
        ws.cell(row=row, column=1, value="Žltá = Neutrálna zóna")
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color=ExportUtils.COLORS['neutral'],
            end_color=ExportUtils.COLORS['neutral'],
            fill_type='solid'
        )
        
        # Šírka stĺpcov
        ws.column_dimensions['A'].width = 10
        for i in range(2, len(price_changes) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 10


def export_strategy(strategy: Dict, scenarios: Dict = None, 
                    alternatives: List[Dict] = None, margin_info: Dict = None,
                    output_dir: str = '/tmp', format: str = 'both') -> Dict:
    """
    Vysokoúrovňová funkcia pre export kompletných výsledkov
    
    Args:
        strategy: Dict so stratégiou (shortLeg, longLeg, strategy)
        scenarios: Dict so scenármi (priceScenarios, timeScenarios, combinedMatrix)
        alternatives: List alternatív s rôznymi DTE
        margin_info: Dict s margin informáciami
        output_dir: Adresár pre výstupné súbory
        format: 'csv', 'excel' alebo 'both'
        
    Returns:
        Dict s cestami k vytvoreným súborom
    """
    # Zostavenie dát pre export
    export_data = {**strategy}
    
    if scenarios:
        if 'priceScenarios' in scenarios:
            export_data['priceScenarios'] = scenarios['priceScenarios']
        elif 'scenarios' in scenarios and 'priceChange' in scenarios.get('scenarios', [{}])[0]:
            export_data['priceScenarios'] = scenarios['scenarios']
        
        if 'timeScenarios' in scenarios:
            export_data['timeScenarios'] = scenarios['timeScenarios']
        
        if 'combinedMatrix' in scenarios:
            export_data['combinedMatrix'] = scenarios['combinedMatrix']
        elif 'matrix' in scenarios:
            export_data['combinedMatrix'] = scenarios
    
    if alternatives:
        export_data['alternatives'] = alternatives
    
    if margin_info:
        export_data['marginInfo'] = margin_info
    
    symbol = strategy.get('symbol', 'UNKNOWN')
    base_filename = ExportUtils.generate_filename(symbol)
    base_path = os.path.join(output_dir, base_filename)
    
    result = {'success': True, 'files': []}
    
    try:
        if format in ['csv', 'both']:
            csv_path = ExportUtils.export_to_csv(export_data, base_path)
            result['files'].append(csv_path)
            result['csvPath'] = csv_path
        
        if format in ['excel', 'both']:
            xlsx_path = ExportUtils.export_to_excel(export_data, base_path)
            result['files'].append(xlsx_path)
            result['excelPath'] = xlsx_path
    
    except Exception as e:
        result['success'] = False
        result['error'] = str(e)
    
    return result


# === TESTY ===
if __name__ == '__main__':
    print("=== TEST EXPORT UTILS ===\n")
    
    # Test data
    strategy = {
        'symbol': 'SPY',
        'currentPrice': 607.50,
        'optionType': 'PUT',
        'shortLeg': {
            'strike': 590,
            'expiry': '20250103',
            'premium': 0.85,
            'delta': -0.0823,
            'theta': -0.0456,
            'iv': 0.18,
        },
        'longLeg': {
            'strike': 565,
            'expiry': '20250117',
            'premium': 0.45,
            'delta': -0.0312,
            'theta': -0.0234,
            'iv': 0.20,
        },
        'strategy': {
            'netCredit': 0.40,
            'maxProfit': 40.00,
            'maxLoss': 2460.00,
            'breakeven': 589.60,
            'spreadWidth': 25,
            'marginRequired': 2500.00,
        }
    }
    
    scenarios = {
        'priceScenarios': [
            {'priceChange': -5, 'newPrice': 577.13, 'pnl': -120.50, 'netDelta': -0.15},
            {'priceChange': -2, 'newPrice': 595.35, 'pnl': -45.20, 'netDelta': -0.08},
            {'priceChange': 0, 'newPrice': 607.50, 'pnl': 5.00, 'netDelta': -0.05},
            {'priceChange': 2, 'newPrice': 619.65, 'pnl': 25.00, 'netDelta': -0.03},
            {'priceChange': 5, 'newPrice': 637.88, 'pnl': 38.00, 'netDelta': -0.01},
        ],
        'timeScenarios': [
            {'daysForward': 0, 'shortDTE': 13, 'longDTE': 27, 'pnl': 5.00},
            {'daysForward': 3, 'shortDTE': 10, 'longDTE': 24, 'pnl': 15.00},
            {'daysForward': 7, 'shortDTE': 6, 'longDTE': 20, 'pnl': 28.00},
        ],
        'combinedMatrix': {
            'currentPrice': 607.50,
            'originalNetCredit': 0.40,
            'priceChanges': [-5, -2, 0, 2, 5],
            'matrix': [
                {'shortDTE': 13, 'scenarios': [
                    {'priceChange': -5, 'pnl': -120, 'zone': 'loss'},
                    {'priceChange': -2, 'pnl': -45, 'zone': 'neutral'},
                    {'priceChange': 0, 'pnl': 5, 'zone': 'neutral'},
                    {'priceChange': 2, 'pnl': 25, 'zone': 'profit'},
                    {'priceChange': 5, 'pnl': 38, 'zone': 'profit'},
                ]},
                {'shortDTE': 6, 'scenarios': [
                    {'priceChange': -5, 'pnl': -80, 'zone': 'neutral'},
                    {'priceChange': -2, 'pnl': -20, 'zone': 'neutral'},
                    {'priceChange': 0, 'pnl': 28, 'zone': 'profit'},
                    {'priceChange': 2, 'pnl': 35, 'zone': 'profit'},
                    {'priceChange': 5, 'pnl': 39, 'zone': 'profit'},
                ]},
            ]
        }
    }
    
    alternatives = [
        {'dteOffset': 0, 'longStrike': 565, 'margin': 2500, 'netCredit': 0.40, 
         'weeklyROI': 1.2, 'thetaAdjustedWeeklyROI': 1.4, 'thetaDifferential': 0.022},
        {'dteOffset': 7, 'longStrike': 565, 'margin': 2800, 'netCredit': 0.30,
         'weeklyROI': 0.8, 'thetaAdjustedWeeklyROI': 1.1, 'thetaDifferential': 0.018},
        {'dteOffset': 14, 'longStrike': 565, 'margin': 3100, 'netCredit': 0.25,
         'weeklyROI': 0.6, 'thetaAdjustedWeeklyROI': 0.9, 'thetaDifferential': 0.015},
    ]
    
    margin_info = {
        'broker': 'IBKR',
        'brokerName': 'Interactive Brokers',
        'spreadType': 'diagonal',
        'margin': 2500,
        'roiOnMargin': 1.6,
        'weeklyROI': 1.2,
        'thetaAdjustedWeeklyROI': 1.4,
    }
    
    # Test export
    result = export_strategy(
        strategy=strategy,
        scenarios=scenarios,
        alternatives=alternatives,
        margin_info=margin_info,
        output_dir='/tmp',
        format='both'
    )
    
    print(f"Export result: {result}")
    
    if result['success']:
        print(f"\nVytvorené súbory:")
        for f in result['files']:
            print(f"  {f}")
